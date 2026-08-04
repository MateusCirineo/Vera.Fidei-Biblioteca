from __future__ import annotations

import io
import csv
import json
import logging
import datetime
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse

from core.deps import get_current_user, require_min_plan
from core.plans import has_feature, history_limit_for_plan
from core.verification_quota import get_verification_usage
from models.database import SessionLocal, User, VerificationHistory
from schemas.citation import VerificationUsage, VerifyCitationRequest, VerifyCitationResponse
from services.verification_service import VerificationService

logger = logging.getLogger(__name__)

router = APIRouter()
service = VerificationService()
BR_TZ = ZoneInfo("America/Sao_Paulo")
CSV_TEXT_LIMIT = 280
CSV_EXCERPT_LIMIT = 360
CSV_ANALYSIS_LIMIT = 360

_FIEL_FIELDS = {
    "reference": None,
    "original_language": None,
    "source_version": None,
    "matched_excerpt": None,
}
_APOLOGETA_FIELDS = {
    "context_before": None,
    "context_after": None,
    "matched_translation": None,
    "translation_language": None,
    "translation_fidelity": None,
    "translator": None,
    "translation_edition": None,
    "variant_analysis": None,
}


def _gate_result_for_plan(result: VerifyCitationResponse, plan: str | None) -> VerifyCitationResponse:
    fields = {}
    if not has_feature(plan, "exact_reference"):
        fields.update(_FIEL_FIELDS)
    if not has_feature(plan, "patristic_context"):
        fields.update(_APOLOGETA_FIELDS)
    return result.model_copy(update=fields) if fields else result


def _current_user_and_quota(current_user: User) -> tuple[User, VerificationUsage]:
    with SessionLocal() as db:
        user = db.get(User, current_user.id)
        if not user or not user.is_active:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Usuário não encontrado ou inativo.")
        quota = get_verification_usage(db, user)
        return user, quota


@router.post("/verify-citation", response_model=VerifyCitationResponse)
def verify_citation(
    payload: VerifyCitationRequest,
    current_user: User = Depends(get_current_user),
) -> VerifyCitationResponse:
    user, quota = _current_user_and_quota(current_user)
    if quota.blocked:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail=jsonable_encoder(
                {
                    "code": "VERIFICATION_LIMIT_REACHED",
                    "message": "Limite mensal de verificações atingido.",
                    "quota": quota,
                }
            ),
        )

    raw_result = service.verify(payload)
    result = _gate_result_for_plan(raw_result, user.plan)

    try:
        ref_json = raw_result.reference.model_dump_json() if raw_result.reference else None
        entry = VerificationHistory(
            user_id=current_user.id,
            citation_text=payload.quote,
            attributed_to=payload.attributed_to,
            status_code=raw_result.status_code,
            label=raw_result.label,
            confidence=raw_result.confidence,
            author=raw_result.author,
            work=raw_result.work,
            reference_json=ref_json,
            matched_excerpt=raw_result.matched_excerpt,
            explanation=raw_result.explanation,
            response_json=raw_result.model_dump_json(),
        )
        with SessionLocal() as db:
            fresh_user = db.get(User, current_user.id)
            db.add(entry)
            db.commit()
            db.refresh(entry)
            updated_quota = get_verification_usage(db, fresh_user) if fresh_user else quota
            result = result.model_copy(update={"history_id": entry.id, "quota": updated_quota})
    except Exception:
        logger.exception("Falha ao salvar histórico para user_id=%s", current_user.id)
        result = result.model_copy(update={"quota": quota})

    return result


@router.get("/usage", response_model=VerificationUsage)
def get_usage(current_user: User = Depends(get_current_user)) -> VerificationUsage:
    _, quota = _current_user_and_quota(current_user)
    return quota


def _load_reference(reference_json: str | None) -> dict | None:
    if not reference_json:
        return None
    try:
        return json.loads(reference_json)
    except Exception:
        return None


def _history_item(
    entry: VerificationHistory,
    *,
    include_reference: bool,
    include_excerpt: bool,
    include_variant: bool,
) -> dict:
    item = {
        "id": entry.id,
        "citation_text": entry.citation_text,
        "attributed_to": entry.attributed_to,
        "status_code": entry.status_code,
        "label": entry.label,
        "confidence": entry.confidence,
        "author": entry.author,
        "work": entry.work,
        "matched_excerpt": entry.matched_excerpt if include_excerpt else None,
        "reference": _load_reference(entry.reference_json) if include_reference else None,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
    }
    if include_variant and entry.response_json:
        try:
            response = json.loads(entry.response_json)
            item["variant_analysis"] = response.get("variant_analysis")
        except Exception:
            item["variant_analysis"] = None
    return item


@router.get("/historico")
def get_historico(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
) -> dict:
    offset = (page - 1) * per_page
    history_cap = history_limit_for_plan(current_user.plan)
    include_reference = has_feature(current_user.plan, "exact_reference")
    include_excerpt = has_feature(current_user.plan, "exact_reference")
    include_variant = has_feature(current_user.plan, "textual_variant_analysis")
    with SessionLocal() as db:
        total_all = db.query(VerificationHistory).filter(
            VerificationHistory.user_id == current_user.id
        ).count()
        total = min(total_all, history_cap) if history_cap is not None else total_all
        effective_limit = per_page
        if history_cap is not None:
            remaining_cap = max(history_cap - offset, 0)
            effective_limit = min(per_page, remaining_cap)
        entries = (
            db.query(VerificationHistory)
            .filter(VerificationHistory.user_id == current_user.id)
            .order_by(VerificationHistory.created_at.desc())
            .offset(offset)
            .limit(effective_limit)
            .all()
            if effective_limit > 0
            else []
        )
        items = [
            _history_item(
                e,
                include_reference=include_reference,
                include_excerpt=include_excerpt,
                include_variant=include_variant,
            )
            for e in entries
        ]
    return {
        "total": total,
        "total_all": total_all,
        "page": page,
        "per_page": per_page,
        "history_scope": "complete" if history_cap is None else "recent",
        "history_limit": history_cap,
        "items": items,
    }


@router.get("/historico/export.csv")
def export_historico_csv(
    current_user: User = Depends(require_min_plan("apologeta")),
) -> StreamingResponse:
    with SessionLocal() as db:
        entries = (
            db.query(VerificationHistory)
            .filter(VerificationHistory.user_id == current_user.id)
            .order_by(VerificationHistory.created_at.desc())
            .all()
        )

    output = io.StringIO()
    output.write("sep=;\r\n")
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")
    writer.writerow([
        "ID",
        "Data da verificação",
        "Citação analisada (resumo)",
        "Atribuída a",
        "Veredito",
        "Confiança",
        "Autor localizado",
        "Obra",
        "Referência",
        "Trecho localizado (resumo)",
        "Análise",
        "Variação textual",
    ])
    for entry in entries:
        ref = _load_reference(entry.reference_json) or {}
        variant = ""
        if entry.response_json:
            try:
                variant = json.loads(entry.response_json).get("variant_analysis") or ""
            except Exception:
                variant = ""
        writer.writerow([
            entry.id,
            _format_csv_datetime(entry.created_at),
            _csv_clean(entry.citation_text, CSV_TEXT_LIMIT),
            entry.attributed_to or "",
            entry.label or entry.status_code or "",
            entry.confidence or "",
            entry.author or "",
            entry.work or "",
            _csv_reference(ref),
            _csv_clean(entry.matched_excerpt or "", CSV_EXCERPT_LIMIT),
            _csv_clean(entry.explanation or "", CSV_ANALYSIS_LIMIT),
            _csv_clean(variant, CSV_ANALYSIS_LIMIT),
        ])

    filename = "historico_verafidei.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/historico/export.xlsx")
def export_historico_xlsx(
    current_user: User = Depends(require_min_plan("apologeta")),
) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    with SessionLocal() as db:
        entries = (
            db.query(VerificationHistory)
            .filter(VerificationHistory.user_id == current_user.id)
            .order_by(VerificationHistory.created_at.desc())
            .all()
        )

    headers = [
        "ID",
        "Data da verificação",
        "Citação analisada",
        "Atribuída a",
        "Veredito",
        "Confiança",
        "Autor localizado",
        "Obra",
        "Referência",
        "Trecho localizado",
        "Análise",
        "Variação textual",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Histórico"
    sheet.append(headers)

    for entry in entries:
        ref = _load_reference(entry.reference_json) or {}
        variant = ""
        if entry.response_json:
            try:
                variant = json.loads(entry.response_json).get("variant_analysis") or ""
            except Exception:
                variant = ""
        sheet.append([
            entry.id,
            _format_csv_datetime(entry.created_at),
            _csv_clean(entry.citation_text, 900),
            entry.attributed_to or "",
            entry.label or entry.status_code or "",
            entry.confidence or "",
            entry.author or "",
            entry.work or "",
            _csv_reference(ref),
            _csv_clean(entry.matched_excerpt or "", 1200),
            _csv_clean(entry.explanation or "", 900),
            _csv_clean(variant, 700),
        ])

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(name="Calibri", color="D9C46A", bold=True, size=11)
    text_font = Font(name="Calibri", color="1F2937", size=10)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = text_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = {
        "A": 8,
        "B": 19,
        "C": 42,
        "D": 24,
        "E": 22,
        "F": 14,
        "G": 26,
        "H": 28,
        "I": 30,
        "J": 52,
        "K": 48,
        "L": 36,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 56

    if sheet.max_row >= 2:
        table = Table(displayName="HistoricoVeraFidei", ref=f"A1:L{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = "historico_verafidei.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_clean(value: str | None, limit: int) -> str:
    text = " ".join((value or "").split())
    if len(text) <= limit:
        return text
    return f"{text[:limit].rstrip()}..."


def _format_csv_datetime(value: datetime.datetime | None) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=datetime.timezone.utc)
    return value.astimezone(BR_TZ).strftime("%d/%m/%Y %H:%M")


def _csv_reference(ref: dict) -> str:
    parts = []
    edition = ref.get("edition_label")
    source = ref.get("source_label")
    page = ref.get("pdf_page")
    section = ref.get("chapter_or_section")
    if edition:
        parts.append(f"Edição: {edition}")
    if source:
        parts.append(f"Fonte: {source}")
    if page:
        parts.append(f"Página PDF: {page}")
    if section:
        parts.append(f"Seção: {section}")
    return " | ".join(parts)


@router.delete("/historico/{entry_id}", status_code=status.HTTP_204_NO_CONTENT, response_model=None)
def delete_historico(
    entry_id: int,
    current_user: User = Depends(get_current_user),
):
    with SessionLocal() as db:
        entry = db.get(VerificationHistory, entry_id)
        if not entry or entry.user_id != current_user.id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entrada não encontrada.")
        db.delete(entry)
        db.commit()


@router.get("/historico/{entry_id}/laudo")
def get_laudo(
    entry_id: int,
    current_user: User = Depends(require_min_plan("catequista")),
) -> StreamingResponse:
    from services.laudo_service import generate_laudo_pdf

    with SessionLocal() as db:
        entry = db.get(VerificationHistory, entry_id)
        if not entry or entry.user_id != current_user.id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entrada não encontrada.")
        pdf_bytes = generate_laudo_pdf(entry, user_plan=current_user.plan)

    filename = f"laudo_verafidei_{entry_id}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/historico/{entry_id}/export-ref")
def export_academic_ref(
    entry_id: int,
    format: str = Query(default="bibtex", pattern="^(bibtex|abnt|ris)$"),
    current_user: User = Depends(require_min_plan("catequista")),
) -> StreamingResponse:
    """Exporta a referência acadêmica de uma verificação em BibTeX, ABNT ou RIS."""
    with SessionLocal() as db:
        entry = db.get(VerificationHistory, entry_id)
        if not entry or entry.user_id != current_user.id:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entrada não encontrada.")
        ref_data = json.loads(entry.reference_json) if entry.reference_json else {}

    author = entry.author or "Autor desconhecido"
    work = entry.work or "Obra desconhecida"
    collection = ref_data.get("collection", "")
    volume = ref_data.get("volume")
    column = ref_data.get("column_start")
    edition_label = ref_data.get("edition_label", "")
    source_label = ref_data.get("source_label", "")
    pdf_page = ref_data.get("pdf_page")
    year = ""

    safe_key = "".join(c for c in (author + work)[:30] if c.isalnum())

    if format == "bibtex":
        lines = [
            f"@book{{{safe_key},",
            f"  author    = {{{author}}},",
            f"  title     = {{{work}}},",
        ]
        if edition_label:
            lines.append(f"  edition   = {{{edition_label}}},")
        if source_label:
            lines.append(f"  publisher = {{{source_label}}},")
        if collection:
            lines.append(f"  series    = {{{collection}}},")
        if volume:
            lines.append(f"  volume    = {{{volume}}},")
        if column:
            lines.append(f"  note      = {{Col. {column}}},")
        if pdf_page:
            lines.append(f"  pages     = {{{pdf_page}}},")
        if year:
            lines.append(f"  year      = {{{year}}},")
        lines.append("}")
        content = "\n".join(lines)
        media = "text/plain"
        filename = f"ref_{entry_id}.bib"

    elif format == "abnt":
        parts = [f"{author.upper()}."]
        parts.append(f" **{work}**.")
        if edition_label:
            parts.append(f" {edition_label}.")
        if source_label:
            parts.append(f" {source_label}.")
        if collection and volume:
            parts.append(f" ({collection}, v. {volume}).")
        elif collection:
            parts.append(f" ({collection}).")
        if column:
            parts.append(f" Col. {column}.")
        content = "".join(parts)
        media = "text/plain"
        filename = f"ref_{entry_id}_abnt.txt"

    else:  # ris
        lines = [
            "TY  - BOOK",
            f"AU  - {author}",
            f"TI  - {work}",
        ]
        if edition_label:
            lines.append(f"ET  - {edition_label}")
        if source_label:
            lines.append(f"PB  - {source_label}")
        if collection:
            lines.append(f"T3  - {collection}")
        if volume:
            lines.append(f"VL  - {volume}")
        if column:
            lines.append(f"SP  - {column}")
        if year:
            lines.append(f"PY  - {year}")
        lines.append("ER  -")
        content = "\n".join(lines)
        media = "application/x-research-info-systems"
        filename = f"ref_{entry_id}.ris"

    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
